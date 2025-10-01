import os
import numpy as np
import polars as pd
from polars import DataFrame as PLDataFrame
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from parser.vital_utils import is_nan, find_latest_vital
from vitaldb import VitalFile
import parser.arr as arr_utils
from zipfile import BadZipFile
from concurrent.futures import ThreadPoolExecutor
import gc


def process_segment(cfg, start_time, vf, arr_utils):
    interval_s = cfg['interval_secs']
    
    # Cargar toda la señal y luego extraer el segmento
    all_data = vf.to_numpy([cfg['signal_track']], interval=0, return_timestamp=True)
    if all_data is None or all_data.shape[0] == 0:
        print(f"[ERROR] Debug: No se pudo cargar señal para {cfg['signal_track']}")
        return None
    
    print(f"[DEBUG] Señal cargada, shape: {all_data.shape}, duración: {all_data[-1, 0] - all_data[0, 0]:.1f}s")

    # Calcular índices para el segmento
    sample_rate = len(all_data) / (all_data[-1, 0] - all_data[0, 0])  # aprox
    start_idx = int(start_time * sample_rate)
    end_idx = int((start_time + interval_s) * sample_rate)
    
    if end_idx > len(all_data):
        return None
    
    # Extraer segmento
    segment_data = all_data[start_idx:end_idx]
    if len(segment_data) == 0:
        return None
        
    timestamps = segment_data[:, 0]
    signal = segment_data[:, 1]
    
    signal = arr_utils.interp_undefined(signal)

    orig_rate = cfg.get('orig_rate', cfg['resample_rate'])
    signal = arr_utils.resample_hz(signal, orig_rate, cfg['resample_rate'])
    timestamps = arr_utils.resample_hz(timestamps, orig_rate, cfg['resample_rate'])

    if len(signal) < cfg['signal_length']:
        print(f"[ERROR] Debug: Señal demasiado corta: {len(signal)} < {cfg['signal_length']}")
        return None

    inp = signal[:cfg['signal_length']].reshape(1, 1, -1)
    try:
        print(f"[DEBUG] Procesando segmento en {start_time}s, señal shape: {inp.shape}")
        pred_result = cfg['model'].predict(inp)
        print(f"[DEBUG] Resultado predicción: {pred_result}")
        pred = float(pred_result.squeeze())
        print(f"[DEBUG] Predicción final: {pred}")
        return {'Tiempo': timestamps[cfg['signal_length'] - 1], cfg['output_var']: pred}
    except Exception as e:
        print(f"[ERROR] Error en process_segment: {e}")
        import traceback
        traceback.print_exc()
        return None

class VitalProcessor:
    def __init__(self, model_configs, results_dir, window_rows=20, poll_interval=1):
        """
        model_configs: lista de dict con configuraciones tabular y wave.
        results_dir: carpeta de salida.
        window_rows: filas para modo tabular.
        poll_interval: no usado aquí.
        """
        self.model_configs = model_configs
        self.results_dir = results_dir
        os.makedirs(self.results_dir, exist_ok=True)
        self.window_rows = window_rows
        self.poll_interval = poll_interval
        self.latest_df = None
        # Estado para limitar predicciones wave según interval_secs
        # clave: output_var, valor: timestamp de última predicción
        self.wave_last = {cfg.get('output_var'): None for cfg in model_configs if cfg.get('input_type')=='wave'}
        # Para rastrear el último tiempo procesado por archivo para sincronización en tiempo real
        self.last_processing_time = {}

    def process_once(self, recordings_dir, mode='tabular'):
        if mode == 'tabular':
            return self._process_tabular(recordings_dir)
        elif mode == 'wave':
            return self._process_wave(recordings_dir)
        else:
            raise ValueError(f"Unknown mode: {mode}")

    def _process_tabular(self, recordings_dir):
        vital_path = find_latest_vital(recordings_dir)
        if not vital_path:
            return None
        base = os.path.splitext(os.path.basename(vital_path))[0]
        xlsx_path = os.path.join(self.results_dir, f"{base}_tabular.xlsx")
        first = not os.path.exists(xlsx_path)

        vf = VitalFile(vital_path)
        tracks = vf.get_track_names()
        raw = vf.to_numpy(tracks, interval=0, return_timestamp=True)
        buffer = []
        for row in raw[-self.window_rows:]:
            t, *vals = row
            if any(not is_nan(v) for v in vals):
                rec = {'Tiempo': t}
                rec.update({n: v for n, v in zip(tracks, vals) if not is_nan(v)})
                buffer.append(rec)
        if not buffer:
            return None

        df = pd.DataFrame(buffer)
        df = self._run_predictions(df)
        self.latest_df = df.clone()
        self._save_excel(df, xlsx_path, first)
        return df

    def _process_wave(self, recordings_dir):
        """
        Procesa pipelines de onda: carga latest.vital, aplica segmentación por ventanas solapadas y modelos wave.
        Une los resultados en un único DataFrame que incluye las señales originales y las predicciones.
        Implementa sincronización en tiempo real para evitar quedarse en el pasado.
        """
        import time
        
        vital_path = find_latest_vital(recordings_dir)
        if not vital_path:
            print("[ERROR] Debug: No se encontró archivo .vital en", recordings_dir)
            return None
        print(f"[DEBUG] Procesando archivo .vital: {vital_path}")
    
        base = os.path.splitext(os.path.basename(vital_path))[0]
        xlsx_path = os.path.join(self.results_dir, f"{base}_wave.xlsx")
        first = not os.path.exists(xlsx_path)
    
        vf = VitalFile(vital_path)
        records = []
        
        # Tiempo de inicio del procesamiento para control de tiempo real
        processing_start_time = time.time()
        
        for cfg in self.model_configs:
            if cfg.get('input_type') != 'wave' or cfg.get('model') is None:
                continue
            
            # Calcular duración para este track específico
            total_duration = len(vf.to_numpy([cfg['signal_track']], interval=1))
            
            interval_s = cfg['interval_secs']
            overlap_s = cfg['overlap_secs']
            step_s = max(1, interval_s - overlap_s)
            
            # Generar todos los start_times posibles
            all_start_times = np.arange(0, total_duration - interval_s, step_s) if total_duration > interval_s else []
            
            if len(all_start_times) == 0:
                continue
            
            # Determinar desde dónde empezar basado en tiempo real y último procesamiento
            current_time = time.time()
            elapsed_processing = current_time - processing_start_time
            
            # Obtener el último tiempo procesado para este archivo
            file_key = os.path.basename(vital_path)
            last_processed = self.last_processing_time.get(file_key, 0)
            
            # Definir batch_size antes de usarlo
            batch_size = 20  # Procesar 20 segmentos por lote
            
            # Lógica de procesamiento por lotes inteligente
            if last_processed > 0:
                # Continuar desde donde se quedó la última vez
                remaining_start_times = all_start_times[all_start_times > last_processed]
                if len(remaining_start_times) > 0:
                    start_times = remaining_start_times
                    print(f"[INFO] Continuando desde {last_processed}s, {len(start_times)} segmentos pendientes")
                else:
                    # Si no hay más segmentos nuevos, procesar los últimos para mantener actualizado
                    start_times = all_start_times[-batch_size:] if len(all_start_times) >= batch_size else all_start_times
                    print(f"[INFO] No hay segmentos nuevos, procesando últimos {len(start_times)} segmentos")
            else:
                # Primera ejecución: procesar desde el inicio
                start_times = all_start_times
                print(f"[INFO] Primera ejecución: procesando desde el inicio ({len(start_times)} segmentos disponibles)")
            
            if len(start_times) == 0:
                continue
            
            # Procesar solo los segmentos necesarios para mantener tiempo real
            num_workers = max(1, os.cpu_count() // 2)
            
            if len(start_times) > batch_size:
                # Procesar solo el siguiente lote de segmentos
                start_times = start_times[:batch_size]
                print(f"[INFO] Procesando lote de {batch_size} segmentos (total disponibles: {len(all_start_times)})")
            else:
                print(f"[INFO] Procesando lote final de {len(start_times)} segmentos")
            
            print(f"[DEBUG] Procesando {len(start_times)} segmentos con {num_workers} workers")
            with ThreadPoolExecutor(max_workers=num_workers) as executor:
                futures = [executor.submit(process_segment, cfg, st, vf, arr_utils) for st in start_times]
                segment_results = [f.result() for f in futures]
            print(f"[DEBUG] {len(segment_results)} resultados obtenidos")
            
            valid_results = [res for res in segment_results if res is not None]
            records.extend(valid_results)
            gc.collect()
            
            # Actualizar el último tiempo procesado para este archivo
            if len(start_times) > 0:
                self.last_processing_time[file_key] = max(start_times)
            
            # Verificar si debemos parar (límite más generoso para procesamiento por lotes)
            current_elapsed = time.time() - processing_start_time
            if current_elapsed > 80.0:  # Si han pasado más de 60 segundos total
                print(f"[INFO] Tiempo límite alcanzado ({current_elapsed:.1f}s), finalizando procesamiento")
                break
    
        if not records:
            return None
    
        # Resto del código igual
        df_preds = pd.DataFrame(records).unique(subset=['Tiempo'], keep='last')
    
        # Cargar unicamente las tracks necesarias para los modelos wave segun model.json
        tracks = [model_config.get('signal_track') for model_config in self.model_configs if model_config.get('input_type') == 'wave']
        raw_all = vf.to_numpy(tracks, interval=0, return_timestamp=True)
        # Separar columnas
        column_data = {}
        for idx, name in enumerate(['Tiempo'] + tracks):
            try:
                col = raw_all[:, idx]
                if name == 'Tiempo':
                    # Mantener Tiempo como float para el join
                    column_data[name] = col
                else:
                    # Convertir otros a strings o None
                    safe_col = [str(val) if val is not None else None for val in col]
                    column_data[name] = safe_col
            except Exception as e:
                print(f"Error procesando columna {name}: {e}")
                column_data[name] = [None] * len(raw_all)

        df_all = pd.DataFrame(column_data)
    
        # Convertir df_preds a dict para merge manual más simple
        pred_dict = {}
        for row in df_preds.to_dicts():
            pred_dict[row['Tiempo']] = row
        
        # Agregar columnas de predicción
        for pred_col in [col for col in df_preds.columns if col != 'Tiempo']:
            df_all = df_all.with_columns(
                pd.col('Tiempo').map_elements(lambda t: pred_dict.get(t, {}).get(pred_col), return_dtype=pd.Float64).alias(pred_col)
            )
        
        df = df_all
    
        self.latest_df = df.clone()
        self._save_excel(df, xlsx_path, first)
        return df


    def _run_predictions(self, df: pd.DataFrame):
        for cfg in self.model_configs:
            if cfg.get('input_type') != 'tabular':
                continue
            # Solo procesar si el modelo está cargado
            if cfg.get('model') is None:
                continue
            ws    = cfg.get('window_size', 1)
            vars_ = cfg.get('input_vars', [])
            cols  = [c for c in vars_ if c in df]
            if not cols:
                continue
            arr = df[cols].fill_null(0).to_numpy()
            preds = []

            if ws == 1:
                for r in arr:
                    try:
                        out = cfg['model'].predict(r.reshape(1, -1))
                        preds.append(float(out.squeeze()))
                    except Exception as e:
                        print(f"Prediction error (tabular): {e}")
                        preds.append(np.nan)
            else:
                pad = np.zeros((max(0, ws - len(arr)), len(cols)))
                win = np.vstack([pad, arr])
                for i in range(len(arr)):
                    chunk = win[i:i+ws].flatten().reshape(1, -1)
                    try:
                        out = cfg['model'].predict(chunk)
                        preds.append(float(out.squeeze()))
                    except Exception as e:
                        print(f"Prediction error (tabular window): {e}")
                        preds.append(np.nan)

            # Agregar columna de predicciones usando sintaxis correcta de Polars
            df = df.with_columns(pd.Series(cfg['output_var'], preds, dtype=pd.Float64))
        return df

    def _save_excel(self, df: pd.DataFrame, path: str, first: bool):
    
        if not isinstance(df, PLDataFrame):
            raise TypeError("Expected a Polars DataFrame")
    
        # Carga o crea libro
        if os.path.exists(path):
            try:
                wb = load_workbook(path)
            except (InvalidFileException, BadZipFile, OSError):
                print(f"[WARNING] Archivo {path} dañado o inválido, se sobrescribirá.")
                wb = Workbook()
        else:
            wb = Workbook()
    
        # Elimina hoja por defecto si existe y está vacía
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
            wb.remove(wb["Sheet"])
    
        # Crea nueva hoja
        ws = wb.create_sheet("Resultado")
    
        # Escribe encabezados
        ws.append(df.columns)
    
        # Escribe filas
        for row in df.iter_rows(named=True):
            ws.append([row[col] for col in df.columns])
        # Guarda archivo
        wb.save(path)